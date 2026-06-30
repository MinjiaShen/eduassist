# =============================================================================
#  excel_writer.py — Excel 追加写入 / 去重 / 处方明细 Sheet
# =============================================================================

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)

# ── 样式常量 ──────────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="2F5496")   # 深蓝
_SUB_FILL      = PatternFill("solid", fgColor="D6E4F7")   # 浅蓝（处方Sheet表头）
_HEADER_FONT   = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_BODY_FONT     = Font(name="微软雅黑", size=10)
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN          = Side(style="thin", color="B0B0B0")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 处方明细 Sheet 的列
HERB_COLUMNS = ["来源文件", "患者姓名", "就诊日期时间", "序号", "药材名", "剂量g"]


def _style_header_row(ws, columns: list[str]):
    """为第1行（表头）应用样式并设列宽。"""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        # 自适应列宽（最窄8，最宽40）
        width = max(8, min(40, len(col_name) * 2 + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _apply_row_style(ws, row_idx: int, n_cols: int, is_alt: bool = False):
    """为数据行应用样式（交替底色）。"""
    fill = PatternFill("solid", fgColor="EEF4FB") if is_alt else None
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font      = _BODY_FONT
        cell.border    = _BORDER
        cell.alignment = _LEFT
        if fill:
            cell.fill = fill


def _get_or_create_workbook() -> openpyxl.Workbook:
    """加载已有 Excel 或新建（含表头）。"""
    output_path = Path(config.OUTPUT_EXCEL)

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        # 检查两个 Sheet 是否存在，不存在则补建
        if config.SHEET_MAIN not in wb.sheetnames:
            ws_main = wb.create_sheet(config.SHEET_MAIN, 0)
            _style_header_row(ws_main, config.MAIN_COLUMNS)
        if config.SHEET_HERBS not in wb.sheetnames:
            ws_herbs = wb.create_sheet(config.SHEET_HERBS)
            _style_header_row(ws_herbs, HERB_COLUMNS)
    else:
        wb = openpyxl.Workbook()
        # 删除默认的空 Sheet（兼容不同 openpyxl 版本/语言）
        default_names = {"Sheet", "工作表"}
        for name in list(wb.sheetnames):
            if name in default_names or name.startswith("Sheet"):
                del wb[name]

        ws_main = wb.create_sheet(config.SHEET_MAIN)
        _style_header_row(ws_main, config.MAIN_COLUMNS)

        ws_herbs = wb.create_sheet(config.SHEET_HERBS)
        _style_header_row(ws_herbs, HERB_COLUMNS)

    return wb


def _build_dedup_set(ws) -> set:
    """
    读取主表已有数据，构建去重键集合。
    去重键 = tuple(DEDUP_KEYS 对应字段的值)
    """
    if ws.max_row < 2:
        return set()

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    key_indices = []
    for key in config.DEDUP_KEYS:
        try:
            key_indices.append(headers.index(key))
        except ValueError:
            key_indices.append(None)

    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        key_vals = tuple(
            (row[i] if i is not None and i < len(row) else None)
            for i in key_indices
        )
        # 全 None 的行不加入去重集合，避免与新记录碰撞
        if not all(v is None for v in key_vals):
            seen.add(key_vals)
    return seen


def _record_dedup_key(record: dict) -> tuple:
    """
    构建去重键。
    当所有键值均为 None 时，返回基于记录唯一标识的元组，
    避免不同患者因 (None, None, None) 碰撞而被误跳过。
    """
    key = tuple(record.get(k) for k in config.DEDUP_KEYS)
    if all(v is None for v in key):
        # 所有去重键为空 → 使用对象 id 保证唯一性（永不跳过）
        return (id(record),)
    return key


def append_record(record: dict) -> bool:
    """
    将一条记录追加写入 Excel。
    返回 True 表示成功写入，False 表示因去重跳过。
    """
    wb = _get_or_create_workbook()
    ws_main  = wb[config.SHEET_MAIN]
    ws_herbs = wb[config.SHEET_HERBS]

    # ── 去重检查 ──────────────────────────────────────────────────────────────
    seen = _build_dedup_set(ws_main)
    key  = _record_dedup_key(record)
    if key in seen:
        logger.info(f"跳过重复记录: {key}")
        wb.close()
        return False

    # ── 写入主表 ──────────────────────────────────────────────────────────────
    next_main_row = ws_main.max_row + 1
    is_alt = (next_main_row % 2 == 0)

    for col_idx, col_name in enumerate(config.MAIN_COLUMNS, start=1):
        value = record.get(col_name)
        ws_main.cell(row=next_main_row, column=col_idx, value=value)

    _apply_row_style(ws_main, next_main_row, len(config.MAIN_COLUMNS), is_alt)

    # ── 写入处方明细 Sheet ────────────────────────────────────────────────────
    herbs = record.get("_herbs") or []
    if herbs:
        next_herb_row = ws_herbs.max_row + 1
        patient_name  = record.get("姓名")
        visit_dt      = record.get("就诊日期时间")
        source_file   = record.get("来源文件")

        for seq, herb in enumerate(herbs, start=1):
            is_alt_h = (next_herb_row % 2 == 0)
            row_data = [
                source_file,
                patient_name,
                visit_dt,
                seq,
                herb.get("name"),
                herb.get("dose_g"),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                ws_herbs.cell(row=next_herb_row, column=col_idx, value=val)
            _apply_row_style(ws_herbs, next_herb_row, len(HERB_COLUMNS), is_alt_h)
            next_herb_row += 1

    # ── 保存 ──────────────────────────────────────────────────────────────────
    wb.save(config.OUTPUT_EXCEL)
    wb.close()
    logger.info(f"已写入: {record.get('姓名')} → {config.OUTPUT_EXCEL}")
    return True



